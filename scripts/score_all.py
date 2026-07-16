"""
score_all.py

Computes P, R, F0.5 (and GLEU for JFLEG) for all six GEC systems across
all three datasets. Reads from outputs.csv (LT + LLMs) and
grammarly_queue.xlsx, then writes results to scores.csv.

Scoring:
  CoNLL-2014 / BEA-2019 : M2 scorer (implemented here, no spaCy needed)
  JFLEG                  : GLEU scorer (from jfleg-master/eval/gleu.py)

Usage:
  python score_all.py

Paths (edit if needed):
  outputs.csv, grammarly_queue.xlsx, and the three dataset folders
  are all expected relative to this script's directory.
"""

import csv
import subprocess
import sys
import tempfile
import os
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

# ── Paths ────────────────────────────────────────────────────────────────
OUTPUTS_CSV      = Path("outputs.csv")
GRAMMARLY_XLSX   = Path("grammarly_queue.xlsx")
SCORES_CSV       = Path("scores.csv")

CONLL_M2         = Path("conll14st-test-data/noalt/official-2014.combined.m2")
BEA_M2           = Path("wi+locness/m2/ABCN.dev.gold.bea19.m2")
JFLEG_SRC        = Path("jfleg-master/test/test.src")
JFLEG_REFS       = [Path(f"jfleg-master/test/test.ref{i}") for i in range(4)]
GLEU_SCRIPT      = Path("jfleg-master/eval/gleu.py")

# ── Schema ───────────────────────────────────────────────────────────────
OUTPUT_FIELDS = ['original_id','dataset','length_bucket','word_count',
                 'sentence','system','model_version','run_index',
                 'corrected','error','timestamp']

SCORE_FIELDS = ['system','dataset','run_index','tp','fp','fn',
                'precision','recall','f05','gleu','num_sentences']


# ═══════════════════════════════════════════════════════════════════════
# 1.  M2 SCORER
#     Implements the standard M2 scoring algorithm without spaCy.
#     Hypothesis edits are derived from token-level difflib alignment.
# ═══════════════════════════════════════════════════════════════════════

def parse_m2_file(m2_path):
    """
    Parses an M2 file into a list of dicts:
      { 'sentence': str (space-tokenised source),
        'edits':    list of (start, end, corrections, annotator_id)
                    where corrections is a list of strings (one per ref) }
    Multiple annotators are merged: a gold edit is accepted if ANY
    annotator marks it (union), which is how the official M2 scorer
    handles CoNLL-2014's two annotators.
    """
    entries = []
    with open(m2_path, encoding='utf-8') as f:
        source = None
        edits_by_ann = defaultdict(list)  # annotator_id -> list of edits
        for line in f:
            line = line.rstrip('\n')
            if line.startswith('S '):
                if source is not None:
                    entries.append(_merge_annotators(source, edits_by_ann))
                source = line[2:]
                edits_by_ann = defaultdict(list)
            elif line.startswith('A '):
                parts = line[2:].split('|||')
                span = parts[0].split()
                start, end = int(span[0]), int(span[1])
                corr = parts[2]
                ann_id = int(parts[5]) if len(parts) > 5 else 0
                # noop edits signal "no correction needed" — keep them
                edits_by_ann[ann_id].append((start, end, corr))
        if source is not None:
            entries.append(_merge_annotators(source, edits_by_ann))
    return entries


def _merge_annotators(source, edits_by_ann):
    """
    Union of edits across annotators (any annotator accepting = gold).
    Returns {'sentence': str, 'gold_edits': set of (start, end, corr)}.
    """
    gold = set()
    has_noop = False
    for ann_id, edits in edits_by_ann.items():
        for (start, end, corr) in edits:
            if start == -1 and end == -1:
                has_noop = True
            else:
                gold.add((start, end, corr))
    return {'sentence': source, 'gold_edits': gold, 'has_noop': has_noop}


def _tokenise(sentence):
    """Split on whitespace — M2 files are already space-tokenised."""
    return sentence.split()


def _compute_edits(src_tokens, hyp_tokens):
    """
    Compute token-level edits (start, end, replacement) between
    src_tokens and hyp_tokens using difflib SequenceMatcher.
    Returns a set of (start, end, replacement_str) triples matching
    the M2 edit format.
    """
    import difflib
    sm = difflib.SequenceMatcher(None, src_tokens, hyp_tokens)
    edits = set()
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == 'equal':
            continue
        replacement = ' '.join(hyp_tokens[j1:j2])  # may be '' for deletions
        edits.add((i1, i2, replacement))
    return edits


def score_m2(gold_entries, hyp_sentences):
    """
    Score a list of hypothesis sentences against gold M2 entries.
    Returns (tp, fp, fn, precision, recall, f05).

    hyp_sentences: list of strings, one per gold entry, same order.
    """
    assert len(gold_entries) == len(hyp_sentences), (
        f"Mismatch: {len(gold_entries)} gold entries vs "
        f"{len(hyp_sentences)} hypotheses"
    )

    tp = fp = fn = 0
    for gold, hyp in zip(gold_entries, hyp_sentences):
        src_tokens = _tokenise(gold['sentence'])
        hyp_tokens = _tokenise(hyp)
        hyp_edits  = _compute_edits(src_tokens, hyp_tokens)
        gold_edits = gold['gold_edits']

        if not gold_edits and gold['has_noop']:
            # Gold says "no correction needed"
            if not hyp_edits:
                tp += 1   # correctly left unchanged
            else:
                fp += len(hyp_edits)  # over-correction
            continue

        # Count matches
        matched = hyp_edits & gold_edits
        tp += len(matched)
        fp += len(hyp_edits - gold_edits)
        fn += len(gold_edits - hyp_edits)

    p, r, f = _f_score(tp, fp, fn, beta=0.5)
    return tp, fp, fn, p, r, f


def _f_score(tp, fp, fn, beta=0.5):
    p = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    r = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    b2 = beta ** 2
    f = (1 + b2) * p * r / (b2 * p + r) if (b2 * p + r) > 0 else 0.0
    return round(p, 4), round(r, 4), round(f, 4)


# ═══════════════════════════════════════════════════════════════════════
# 2.  GLEU SCORER  (delegates to the official script)
# ═══════════════════════════════════════════════════════════════════════

def score_gleu(sys_run_data, src_path, ref_paths, gleu_script):
    """
    Builds a full hypothesis file (same length as source) for GLEU scoring.
    Non-sampled positions are filled with the original source sentence
    (= no correction). Returns the corpus-level GLEU score (float).
    """
    src_lines = open(src_path, encoding='utf-8').read().splitlines()
    hyp_lines = []
    for i, src in enumerate(src_lines):
        oid = f'jfleg_{i}'
        hyp = sys_run_data.get(oid, src)
        # Sanitize: if LLM returned preamble + newline + sentence,
        # take the last non-empty line as the actual correction.
        if '\n' in hyp:
            parts = [p.strip() for p in hyp.split('\n') if p.strip()]
            hyp = parts[-1] if parts else src
        hyp_lines.append(hyp)

    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt',
                                     delete=False, encoding='utf-8') as tmp:
        tmp.write('\n'.join(hyp_lines) + '\n')
        tmp_path = tmp.name
    try:
        cmd = [sys.executable, str(gleu_script),
               '-r'] + [str(r) for r in ref_paths] + \
              ['-s', str(src_path), '--hyp', tmp_path]
        result = subprocess.run(cmd, capture_output=True, text=True, check=True)
        import ast
        parsed = ast.literal_eval(result.stdout.strip().split('\n')[-1])
        return round(float(parsed[0][0]), 6)
    finally:
        os.unlink(tmp_path)


# ═══════════════════════════════════════════════════════════════════════
# 3.  DATA LOADING
# ═══════════════════════════════════════════════════════════════════════

def load_outputs_csv(csv_path):
    """
    Loads outputs.csv (no header) into a nested dict:
      data[system][dataset][run_index] -> {original_id -> corrected}
    """
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    with open(csv_path, encoding='utf-8') as f:
        for row in csv.DictReader(f, fieldnames=OUTPUT_FIELDS):
            sys_  = row['system']
            ds    = row['dataset']
            run   = row['run_index']
            oid   = row['original_id']
            corr  = row['corrected']
            data[sys_][ds][run][oid] = corr
    return data


def load_grammarly(xlsx_path):
    """
    Reads grammarly_queue.xlsx and extracts the final corrected sentence
    for each row (last non-empty Iteration_N column).
    Returns dict: original_id -> corrected_sentence
    """
    wb = load_workbook(xlsx_path)
    ws = wb['Grammarly_Queue']

    # Iteration text columns: F(6), H(8), J(10), L(12), N(14)
    iter_text_cols = [6, 8, 10, 12, 14]
    result = {}

    for row_idx in range(2, ws.max_row + 1):
        orig_id = ws.cell(row=row_idx, column=2).value
        if not orig_id:
            continue
        last_text = None
        for col in iter_text_cols:
            val = ws.cell(row=row_idx, column=col).value
            if val and str(val).strip():
                last_text = str(val).strip()
        if last_text:
            result[orig_id] = last_text
    return result


def load_m2_indexed(m2_path, dataset_prefix):
    """
    Parses M2 file and returns dict: index (int) -> gold_entry
    so we can look up by original_id (e.g. 'conll14_855' -> index 855).
    """
    entries = parse_m2_file(m2_path)
    return {i: e for i, e in enumerate(entries)}


def load_jfleg_src(src_path):
    """Returns dict: index (int) -> source sentence string."""
    with open(src_path, encoding='utf-8') as f:
        lines = [l.rstrip('\n') for l in f if l.strip()]
    return {i: s for i, s in enumerate(lines)}


# ═══════════════════════════════════════════════════════════════════════
# 4.  SCORING PIPELINE
# ═══════════════════════════════════════════════════════════════════════

def get_ordered_hypotheses(system_run_data, gold_index, dataset_prefix):
    """
    Given system_run_data (original_id -> corrected) and gold_index
    (int -> gold entry), returns a list of (gold_entry, hyp_sentence)
    in the same order as the gold index, for only the sampled sentences.
    """
    # Identify which indices we have (from the original_ids in the data)
    pairs = []
    for orig_id, hyp in system_run_data.items():
        idx = int(orig_id.split('_')[1])
        if idx in gold_index:
            pairs.append((idx, gold_index[idx], hyp))
    # Sort by gold index to ensure consistent ordering
    pairs.sort(key=lambda x: x[0])
    return [(g, h) for _, g, h in pairs]


def score_system_dataset(system, run, dataset, sys_run_data,
                          gold_m2_index=None, jfleg_src=None):
    """
    Score one (system, run, dataset) combination.
    Returns dict of score fields.
    """
    dataset_prefix = dataset  # e.g. 'conll14', 'bea19', 'jfleg'

    if dataset == 'jfleg':
        # For JFLEG: align by index, build ordered hyp list
        gleu = score_gleu(sys_run_data, JFLEG_SRC, JFLEG_REFS, GLEU_SCRIPT)
        return {
            'system': system, 'dataset': dataset, 'run_index': run,
            'tp': '', 'fp': '', 'fn': '',
            'precision': '', 'recall': '', 'f05': '',
            'gleu': gleu, 'num_sentences': len(sys_run_data),
        }
    else:
        # M2 scoring for CoNLL-2014 and BEA-2019
        pairs = get_ordered_hypotheses(sys_run_data, gold_m2_index, dataset_prefix)
        gold_entries = [g for g, _ in pairs]
        hyp_sentences = [h for _, h in pairs]

        tp, fp, fn, p, r, f = score_m2(gold_entries, hyp_sentences)
        return {
            'system': system, 'dataset': dataset, 'run_index': run,
            'tp': tp, 'fp': fp, 'fn': fn,
            'precision': p, 'recall': r, 'f05': f,
            'gleu': '', 'num_sentences': len(hyp_sentences),
        }


# ═══════════════════════════════════════════════════════════════════════
# 5.  MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    print("Loading data...")
    outputs   = load_outputs_csv(OUTPUTS_CSV)
    grammarly = load_grammarly(GRAMMARLY_XLSX)

    print("Parsing gold references...")
    conll_gold  = load_m2_indexed(CONLL_M2, 'conll14')
    bea_gold    = load_m2_indexed(BEA_M2, 'bea19')
    jfleg_src   = load_jfleg_src(JFLEG_SRC)

    gold = {
        'conll14': conll_gold,
        'bea19':   bea_gold,
        'jfleg':   jfleg_src,
    }

    # Add Grammarly to the outputs structure
    # Grammarly: CoNLL-2014 only, 1 run, model_version fixed
    outputs['grammarly']['conll14']['1'] = grammarly

    all_scores = []

    systems = sorted(outputs.keys())
    for system in systems:
        datasets = sorted(outputs[system].keys())
        for dataset in datasets:
            runs = sorted(outputs[system][dataset].keys())
            for run in runs:
                sys_run_data = outputs[system][dataset][run]
                print(f"  Scoring {system:12s} | {dataset:8s} | run {run} "
                      f"({len(sys_run_data)} sentences)...", end=' ', flush=True)

                try:
                    score = score_system_dataset(
                        system, run, dataset, sys_run_data,
                        gold_m2_index=gold.get(dataset),
                        jfleg_src=jfleg_src if dataset == 'jfleg' else None,
                    )
                    all_scores.append(score)
                    if dataset == 'jfleg':
                        print(f"GLEU={score['gleu']:.4f}")
                    else:
                        print(f"F0.5={score['f05']:.4f} "
                              f"P={score['precision']:.4f} "
                              f"R={score['recall']:.4f}")
                except Exception as e:
                    print(f"ERROR: {e}")
                    import traceback; traceback.print_exc()

    # Write results
    with open(SCORES_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=SCORE_FIELDS)
        writer.writeheader()
        writer.writerows(all_scores)

    print(f"\nSaved {len(all_scores)} score rows to {SCORES_CSV}")


if __name__ == '__main__':
    main()
